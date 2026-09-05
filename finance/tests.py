from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from customers.models import Customer
from products.models import Product, Purchase
from sales.models import Sale
from suppliers.models import Supplier
from .models import CashExpense, CashSession, PaymentMethod


class CashValidationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='cashier', password='pass12345', is_staff=True)
        self.client.force_login(self.user)

    def test_open_cash_rejects_negative_initial_amount(self):
        self.client.post(reverse('finance:open_cash'), {'amount': '-100.00'})

        self.assertEqual(CashSession.objects.count(), 0)

    def test_expense_rejects_non_positive_amount(self):
        CashSession.objects.create(user=self.user, initial_amount=Decimal('100.00'))

        self.client.post(reverse('finance:add_expense'), {
            'amount': '-1.00',
            'description': 'Ajuste invalido',
        })

        self.assertEqual(CashExpense.objects.count(), 0)

    def test_cash_report_pdf_exports_without_payment_method_key_error(self):
        session = CashSession.objects.create(
            user=self.user,
            initial_amount=Decimal('100.00'),
            expected_final_amount=Decimal('200.00'),
        )
        cash = PaymentMethod.objects.create(name='Efectivo')
        card = PaymentMethod.objects.create(name='Tarjeta', is_digital=True)
        customer = Customer.objects.create(full_name='Cliente', dni_cuit='99887766')
        Sale.objects.create(user=self.user, customer=customer, total_amount=Decimal('100.00'), payment_method=cash)
        Sale.objects.create(user=self.user, customer=customer, total_amount=Decimal('50.00'), payment_method=card)

        response = self.client.get(reverse('finance:export_cash_report', args=[session.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(bytes(response.content).startswith(b'%PDF'))

    def test_fiscal_report_exports_valid_excel(self):
        supplier = Supplier.objects.create(name='Proveedor Fiscal')
        Purchase.objects.create(supplier=supplier, total_amount=Decimal('121.00'), user=self.user)

        response = self.client.get(reverse('finance:fiscal_reports'), {'export': '1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn('IVA Ventas', workbook.sheetnames)
        self.assertIn('IVA Compras', workbook.sheetnames)
