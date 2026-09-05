import json
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from suppliers.models import Supplier
from .models import Product, Purchase


class ProductStockValidationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='operator', password='pass12345', is_staff=True)
        self.client.force_login(self.user)
        self.supplier = Supplier.objects.create(name='Proveedor')
        self.product = Product.objects.create(
            sku='PROD1',
            barcode='B001',
            name='Producto',
            price=Decimal('50.00'),
            cost_price=Decimal('25.00'),
            stock=10,
            min_stock=2,
            supplier=self.supplier,
        )

    def test_stock_loss_rejects_negative_quantity(self):
        self.client.post(reverse('products:stock_loss_create'), {
            'product': str(self.product.id),
            'quantity': '-5',
            'reason': 'OTH',
        })

        self.product.refresh_from_db()
        self.assertEqual(self.product.stock, 10)

    def test_purchase_rejects_negative_quantity(self):
        items = [{'product_id': self.product.id, 'qty': '-5', 'cost': '10.00'}]

        self.client.post(reverse('products:purchase_create'), {
            'supplier_id': str(self.supplier.id),
            'invoice_number': 'A-1',
            'items_data': json.dumps(items),
        })

        self.assertEqual(Purchase.objects.count(), 0)
        self.product.refresh_from_db()
        self.assertEqual(self.product.stock, 10)

    def test_inventory_excel_exports_valid_workbook_with_barcode_columns(self):
        response = self.client.get(reverse('products:export_inventory_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn('sku', headers)
        self.assertIn('barcode', headers)

    def test_advanced_excel_exports_valid_workbook_even_without_data(self):
        Product.objects.all().delete()

        response = self.client.get(reverse('products:export_advanced_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertGreaterEqual(len(workbook.sheetnames), 1)

    def test_label_export_returns_pdf_for_barcode_product(self):
        response = self.client.get(reverse('products:export_labels'), {
            'products': [str(self.product.id)],
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
